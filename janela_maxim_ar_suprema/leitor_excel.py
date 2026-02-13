"""
Leitor da planilha Excel da Janela Maxim-Ar Suprema.
Parseia as 3 secoes (QUADRO, FOLHA, VIDRO) e retorna DadosJanela.
"""
import openpyxl
from modelos import Peca, DadosJanela, TipoSecao, PosicaoPeca


# Mapeamento de linha -> posicao (mesma ordem em cada secao)
POSICAO_POR_OFFSET = {
    0: PosicaoPeca.LARGURA_S,
    1: PosicaoPeca.LARGURA_I,
    2: PosicaoPeca.ALTURA_E,
    3: PosicaoPeca.ALTURA_D,
}

# Linhas iniciais de cada secao
LINHA_QUADRO = 4
LINHA_FOLHA = 11
LINHA_VIDRO = 18


def _ler_peca(ws, linha, secao, posicao):
    """Le uma peca de uma linha da planilha."""
    perfil = ws[f'L{linha}'].value or ws[f'F{linha}'].value or ""
    perfil = str(perfil).strip()
    if not perfil and secao == TipoSecao.VIDRO:
        perfil = "Vidro"

    corte_esq = ws[f'M{linha}'].value
    corte_dir = ws[f'N{linha}'].value
    medida_com = ws[f'O{linha}'].value
    medida_sem = ws[f'Q{linha}'].value
    desconto_com = ws[f'G{linha}'].value or 0
    desconto_sem = ws[f'H{linha}'].value or 0
    formula = ws[f'E{linha}'].value or ""

    return Peca(
        secao=secao,
        posicao=posicao,
        perfil=perfil,
        medida_com=float(medida_com) if medida_com is not None else 0.0,
        medida_sem=float(medida_sem) if medida_sem is not None else 0.0,
        corte_esq=int(corte_esq) if corte_esq is not None else 90,
        corte_dir=int(corte_dir) if corte_dir is not None else 90,
        desconto_com=float(desconto_com),
        desconto_sem=float(desconto_sem),
        formula=str(formula),
    )


def _ler_secao(ws, linha_inicial, secao):
    """Le as 4 pecas de uma secao (QUADRO, FOLHA ou VIDRO)."""
    pecas = []
    for offset in range(4):
        linha = linha_inicial + offset
        posicao = POSICAO_POR_OFFSET[offset]
        pecas.append(_ler_peca(ws, linha, secao, posicao))
    return pecas


def ler_excel(caminho: str) -> DadosJanela:
    """Le a planilha e retorna os dados completos da janela."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active

    largura_vao = float(ws['B4'].value)
    altura_vao = float(ws['C4'].value)

    pecas_quadro = _ler_secao(ws, LINHA_QUADRO, TipoSecao.QUADRO)
    pecas_folha = _ler_secao(ws, LINHA_FOLHA, TipoSecao.FOLHA)
    pecas_vidro = _ler_secao(ws, LINHA_VIDRO, TipoSecao.VIDRO)

    wb.close()

    return DadosJanela(
        largura_vao=largura_vao,
        altura_vao=altura_vao,
        pecas_quadro=pecas_quadro,
        pecas_folha=pecas_folha,
        pecas_vidro=pecas_vidro,
    )


if __name__ == "__main__":
    import os
    caminho = os.path.join(os.path.dirname(__file__), "janela_maxim_ar.xlsx")
    dados = ler_excel(caminho)
    print(f"Vao: {dados.largura_vao} x {dados.altura_vao} mm")
    print()
    for secao_nome, pecas in [("QUADRO", dados.pecas_quadro),
                               ("FOLHA", dados.pecas_folha),
                               ("VIDRO", dados.pecas_vidro)]:
        print(f"--- {secao_nome} ---")
        for p in pecas:
            print(f"  {p.posicao.value:12s} | {p.perfil:6s} | "
                  f"COM={p.medida_com:8.1f} | SEM={p.medida_sem:8.1f} | "
                  f"Corte {p.corte_esq}/{p.corte_dir}")
        print()
